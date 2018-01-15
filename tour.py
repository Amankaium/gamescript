from openpyxl import load_workbook, Workbook
import datetime

######################
# Актуализация списка сотрудников

# ссылка
url = 'test.xlsx'

# файл
file = load_workbook(filename = url)

# листы файла
fileSheets = file.sheetnames

# Градация баллов
pointsSheet = file['ТП']

pointsDict = {}

for i in range(2, len(pointsSheet['A']) + 1):
    tp = pointsSheet['A%d' % i].value
    p = pointsSheet['C%d' % i].value
    pointsDict.setdefault(tp, p)

# Сотрудники (с 0 баллами по умолчанию)
workersSheet = file['сотрудники']

workersDict = {}

for i in range(2, len(workersSheet['C']) + 1):
    dvo = workersSheet['C%d' % i].value
    fio = workersSheet['F%d' % i].value
    workersDict.setdefault(dvo, [fio, points])
    for tp in pointsDict:




    


# Операции (самое долгое)
operatinos = load_workbook(filename = 'fortest.xlsx')
operationsSheet = operatinos['исх']

for i in range(2, len(operationsSheet['AA']) + 1):
    dvo = operationsSheet['AA%d' % i].value
    tp = operationsSheet['H%d' % i].value    
    if dvo in workersDict and tp in pointsDict:
        workersDict[dvo][1] += pointsDict[tp]


        
'''
получается, что словарь workers становится консалидированным 
списком прошлых операций, теперь надо извлечь прошлый бой, внести
изменения и записать новый бой
'''

boiSheet= file['прошлый бой']

boiDict = {}

j = 1
cells = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

boiDict.setdefault(j, [])
for c in cells:
        boiDict[j].append(boiSheet['%s%d' % (c, 1)].value)

for i in range(2, len(boiSheet['B']) + 1):
    j += 1
    boiDict.setdefault(j, [])
    
    for c in cells:
        boiDict[j].append(boiSheet['%s%d' % (c, i)].value)
    
    for dvo in workersDict:
        if workersDict[dvo][0] == boiDict[j][0]:
            boiDict[j][3] += workersDict[dvo][1]
        elif workersDict[dvo][0] == boiDict[j][5]:
            boiDict[j][8] += workersDict[dvo][1]


# осталось словарь перенести в новый файл
boiNew = openpyxl.Workbook()
boiNewSheet = boiNew.active

for i in range(1, len(boiDict) + 1):    
    boiNewSheet.append(boiDict[i])

today = datetime.datetime.today().strftime('%d.%m.%Y')
        
boiNew.save(today +'.xlsx')
            
        







  


        

